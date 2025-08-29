"""
Advanced report generators for analytics data.
Supports PDF, Excel, and automated business intelligence reports.
"""

import io
import json
import logging
from dataclasses import dataclass
from datetime import datetime
from datetime import timedelta
from typing import Any

from django.utils import timezone

from aura.analytics import get_events
from aura.analytics.models import AlertInstance
from aura.analytics.models import MetricsSnapshot

logger = logging.getLogger(__name__)


@dataclass
class ReportConfig:
    """Configuration for report generation."""

    title: str
    description: str
    period_start: datetime
    period_end: datetime
    include_charts: bool = True
    include_raw_data: bool = False
    format: str = "pdf"  # pdf, excel, json
    sections: list[str] = None


class BaseReportGenerator:
    """Base class for all report generators."""

    def __init__(self, config: ReportConfig):
        self.config = config
        self.data = {}

    def generate(self) -> dict[str, Any]:
        """Generate the report and return metadata."""
        logger.info(f"Generating {self.config.format} report: {self.config.title}")

        try:
            # Collect data
            self._collect_data()

            # Generate report based on format
            if self.config.format == "pdf":
                content, filename = self._generate_pdf()
            elif self.config.format == "excel":
                content, filename = self._generate_excel()
            elif self.config.format == "json":
                content, filename = self._generate_json()
            else:
                raise ValueError(f"Unsupported format: {self.config.format}")

            return {
                "content": content,
                "filename": filename,
                "format": self.config.format,
                "size": len(content)
                if isinstance(content, bytes)
                else len(str(content)),
                "generated_at": timezone.now(),
                "config": self.config,
            }

        except Exception as e:
            logger.error(f"Failed to generate report: {e}")
            raise

    def _collect_data(self):
        """Collect data for the report. Override in subclasses."""

    def _generate_pdf(self) -> tuple[bytes, str]:
        """Generate PDF report."""
        try:
            from django.template.loader import render_to_string
            from weasyprint import CSS
            from weasyprint import HTML

            # Render HTML template
            html_content = render_to_string(
                "analytics/reports/base_report.html",
                {
                    "config": self.config,
                    "data": self.data,
                    "generated_at": timezone.now(),
                },
            )

            # Generate PDF
            pdf_file = HTML(string=html_content).write_pdf()
            filename = f"{self.config.title.lower().replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"

            return pdf_file, filename

        except ImportError:
            logger.error("WeasyPrint not installed. Cannot generate PDF reports.")
            raise

    def _generate_excel(self) -> tuple[bytes, str]:
        """Generate Excel report."""
        try:
            import openpyxl
            from openpyxl.chart import BarChart
            from openpyxl.chart import LineChart
            from openpyxl.chart import Reference
            from openpyxl.styles import Alignment
            from openpyxl.styles import Font
            from openpyxl.styles import PatternFill

            # Create workbook
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Add summary sheet
            self._add_summary_sheet(wb)

            # Add data sheets
            self._add_data_sheets(wb)

            # Add charts if requested
            if self.config.include_charts:
                self._add_chart_sheets(wb)

            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_content = excel_buffer.getvalue()
            excel_buffer.close()

            filename = f"{self.config.title.lower().replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            return excel_content, filename

        except ImportError:
            logger.error("openpyxl not installed. Cannot generate Excel reports.")
            raise

    def _generate_json(self) -> tuple[str, str]:
        """Generate JSON report."""
        report_data = {
            "metadata": {
                "title": self.config.title,
                "description": self.config.description,
                "period_start": self.config.period_start.isoformat(),
                "period_end": self.config.period_end.isoformat(),
                "generated_at": timezone.now().isoformat(),
            },
            "data": self.data,
        }

        json_content = json.dumps(report_data, indent=2, default=str)
        filename = f"{self.config.title.lower().replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"

        return json_content, filename

    def _add_summary_sheet(self, workbook):
        """Add summary sheet to Excel workbook."""
        from openpyxl.styles import Font

        ws = workbook.create_sheet("Summary")

        # Title
        ws["A1"] = self.config.title
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        # Period
        ws["A3"] = "Report Period:"
        ws["B3"] = (
            f"{self.config.period_start.strftime('%Y-%m-%d')} to {self.config.period_end.strftime('%Y-%m-%d')}"
        )

        # Key metrics
        row = 5
        for key, value in self.data.get("summary", {}).items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            ws[f"B{row}"] = value
            row += 1

    def _add_data_sheets(self, workbook):
        """Add data sheets to Excel workbook."""
        # Override in subclasses

    def _add_chart_sheets(self, workbook):
        """Add chart sheets to Excel workbook."""
        # Override in subclasses


class AnalyticsSummaryReportGenerator(BaseReportGenerator):
    """Generator for comprehensive analytics summary reports."""

    def _collect_data(self):
        """Collect comprehensive analytics data."""
        # Get events for the period
        events = get_events(
            start_time=self.config.period_start,
            end_time=self.config.period_end,
            limit=50000,
        )

        # Get metrics snapshots
        snapshots = MetricsSnapshot.objects.filter(
            period_start__gte=self.config.period_start,
            period_end__lte=self.config.period_end,
        ).order_by("period_start")

        # Get alert instances
        alerts = AlertInstance.objects.filter(
            created_at__gte=self.config.period_start,
            created_at__lte=self.config.period_end,
        ).order_by("-created_at")

        # Process data
        self.data = {
            "summary": self._calculate_summary_metrics(events, snapshots, alerts),
            "events": self._process_events_data(events),
            "snapshots": self._process_snapshots_data(snapshots),
            "alerts": self._process_alerts_data(alerts),
            "trends": self._calculate_trends(snapshots),
            "top_events": self._get_top_events(events),
            "user_activity": self._analyze_user_activity(events),
            "system_performance": self._analyze_system_performance(snapshots),
        }

    def _calculate_summary_metrics(self, events, snapshots, alerts):
        """Calculate summary metrics."""
        total_events = len(events)
        unique_users = len(set(e.get("user_id") for e in events if e.get("user_id")))
        total_alerts = alerts.count()
        critical_alerts = alerts.filter(severity="critical").count()

        # Calculate period duration
        duration_days = (self.config.period_end - self.config.period_start).days

        return {
            "total_events": total_events,
            "unique_users": unique_users,
            "events_per_day": total_events / max(1, duration_days),
            "total_alerts": total_alerts,
            "critical_alerts": critical_alerts,
            "alert_rate": (total_alerts / max(1, total_events)) * 100,
            "data_quality_score": self._calculate_data_quality(events),
            "period_duration_days": duration_days,
        }

    def _process_events_data(self, events):
        """Process events data for reporting."""
        # Group by event type
        event_types = {}
        hourly_distribution = {}
        daily_distribution = {}

        for event in events:
            event_type = event.get("event_type", "unknown")
            event_types[event_type] = event_types.get(event_type, 0) + 1

            # Parse timestamp for distributions
            timestamp_str = event.get("timestamp", "")
            if timestamp_str:
                try:
                    if isinstance(timestamp_str, str):
                        dt = datetime.fromisoformat(
                            timestamp_str.replace("Z", "+00:00"),
                        )
                    else:
                        dt = timestamp_str

                    # Hourly distribution
                    hour_key = dt.strftime("%H:00")
                    hourly_distribution[hour_key] = (
                        hourly_distribution.get(hour_key, 0) + 1
                    )

                    # Daily distribution
                    day_key = dt.strftime("%Y-%m-%d")
                    daily_distribution[day_key] = daily_distribution.get(day_key, 0) + 1

                except (ValueError, AttributeError):
                    continue

        return {
            "by_type": event_types,
            "hourly_distribution": hourly_distribution,
            "daily_distribution": daily_distribution,
            "total_count": len(events),
        }

    def _process_snapshots_data(self, snapshots):
        """Process metrics snapshots data."""
        snapshots_data = []

        for snapshot in snapshots:
            snapshots_data.append(
                {
                    "period_start": snapshot.period_start,
                    "period_end": snapshot.period_end,
                    "aggregation_type": snapshot.aggregation_type,
                    "total_events": snapshot.total_events,
                    "unique_users": snapshot.unique_users,
                    "top_event_type": snapshot.top_event_type,
                    "data_quality_score": snapshot.data_quality_score,
                },
            )

        return snapshots_data

    def _process_alerts_data(self, alerts):
        """Process alerts data."""
        alerts_data = []
        severity_counts = {"critical": 0, "error": 0, "warning": 0, "info": 0}
        status_counts = {"active": 0, "acknowledged": 0, "resolved": 0, "dismissed": 0}

        for alert in alerts:
            alerts_data.append(
                {
                    "rule_name": alert.rule.name,
                    "severity": alert.severity,
                    "status": alert.status,
                    "triggered_value": alert.triggered_value,
                    "threshold_value": alert.threshold_value,
                    "created_at": alert.created_at,
                    "acknowledged_at": alert.acknowledged_at,
                    "resolved_at": alert.resolved_at,
                },
            )

            severity_counts[alert.severity] = severity_counts.get(alert.severity, 0) + 1
            status_counts[alert.status] = status_counts.get(alert.status, 0) + 1

        return {
            "alerts": alerts_data,
            "severity_counts": severity_counts,
            "status_counts": status_counts,
            "total_count": len(alerts_data),
        }

    def _calculate_trends(self, snapshots):
        """Calculate trends from snapshots."""
        if len(snapshots) < 2:
            return {}

        # Get first and last snapshots for comparison
        first = snapshots.first()
        last = snapshots.last()

        if not first or not last:
            return {}

        # Calculate percentage changes
        trends = {}

        if first.total_events > 0:
            events_change = (
                (last.total_events - first.total_events) / first.total_events
            ) * 100
            trends["events_trend"] = round(events_change, 2)

        if first.unique_users > 0:
            users_change = (
                (last.unique_users - first.unique_users) / first.unique_users
            ) * 100
            trends["users_trend"] = round(users_change, 2)

        # Calculate average values
        total_events = sum(s.total_events for s in snapshots)
        total_users = sum(s.unique_users for s in snapshots)

        trends["avg_events_per_period"] = round(total_events / len(snapshots), 2)
        trends["avg_users_per_period"] = round(total_users / len(snapshots), 2)

        return trends

    def _get_top_events(self, events, limit=10):
        """Get top event types."""
        event_counts = {}

        for event in events:
            event_type = event.get("event_type", "unknown")
            event_counts[event_type] = event_counts.get(event_type, 0) + 1

        # Sort by count and return top N
        sorted_events = sorted(event_counts.items(), key=lambda x: x[1], reverse=True)
        return sorted_events[:limit]

    def _analyze_user_activity(self, events):
        """Analyze user activity patterns."""
        user_events = {}
        user_sessions = {}

        for event in events:
            user_id = event.get("user_id")
            if not user_id:
                continue

            # Count events per user
            user_events[user_id] = user_events.get(user_id, 0) + 1

            # Track sessions (simplified - same day = same session)
            timestamp_str = event.get("timestamp", "")
            if timestamp_str:
                try:
                    if isinstance(timestamp_str, str):
                        dt = datetime.fromisoformat(
                            timestamp_str.replace("Z", "+00:00"),
                        )
                    else:
                        dt = timestamp_str

                    day_key = dt.strftime("%Y-%m-%d")
                    session_key = f"{user_id}_{day_key}"
                    user_sessions[session_key] = True

                except (ValueError, AttributeError):
                    continue

        # Calculate statistics
        if user_events:
            event_counts = list(user_events.values())
            avg_events_per_user = sum(event_counts) / len(event_counts)
            max_events_per_user = max(event_counts)
            min_events_per_user = min(event_counts)
        else:
            avg_events_per_user = max_events_per_user = min_events_per_user = 0

        return {
            "total_active_users": len(user_events),
            "total_sessions": len(user_sessions),
            "avg_events_per_user": round(avg_events_per_user, 2),
            "max_events_per_user": max_events_per_user,
            "min_events_per_user": min_events_per_user,
            "avg_sessions_per_user": round(
                len(user_sessions) / max(1, len(user_events)),
                2,
            ),
        }

    def _analyze_system_performance(self, snapshots):
        """Analyze system performance from snapshots."""
        if not snapshots:
            return {}

        # Calculate performance metrics
        total_events = sum(s.total_events for s in snapshots)
        avg_events_per_period = total_events / len(snapshots)

        # Data quality analysis
        quality_scores = [
            s.data_quality_score for s in snapshots if s.data_quality_score
        ]
        avg_quality = sum(quality_scores) / len(quality_scores) if quality_scores else 0

        # Peak analysis
        max_events = max(s.total_events for s in snapshots)
        min_events = min(s.total_events for s in snapshots)

        return {
            "avg_events_per_period": round(avg_events_per_period, 2),
            "peak_events": max_events,
            "minimum_events": min_events,
            "avg_data_quality": round(avg_quality, 3),
            "total_periods_analyzed": len(snapshots),
            "performance_stability": round(
                (1 - (max_events - min_events) / max(1, avg_events_per_period)) * 100,
                2,
            ),
        }

    def _calculate_data_quality(self, events):
        """Calculate overall data quality score."""
        if not events:
            return 0.0

        # Check for required fields
        complete_events = 0
        for event in events:
            score = 0
            if event.get("event_type"):
                score += 0.3
            if event.get("timestamp"):
                score += 0.3
            if event.get("user_id"):
                score += 0.2
            if event.get("data"):
                score += 0.2

            if score >= 0.8:  # 80% complete
                complete_events += 1

        return round((complete_events / len(events)) * 100, 2)


class BusinessIntelligenceReportGenerator(BaseReportGenerator):
    """Generator for business intelligence reports with healthcare-specific insights."""

    def _collect_data(self):
        """Collect BI-specific data."""
        events = get_events(
            start_time=self.config.period_start,
            end_time=self.config.period_end,
            limit=50000,
        )

        self.data = {
            "patient_insights": self._analyze_patient_metrics(events),
            "therapy_insights": self._analyze_therapy_metrics(events),
            "communication_insights": self._analyze_communication_metrics(events),
            "operational_insights": self._analyze_operational_metrics(events),
            "revenue_insights": self._analyze_revenue_impact(events),
            "recommendations": self._generate_recommendations(events),
        }

    def _analyze_patient_metrics(self, events):
        """Analyze patient-related metrics."""
        patient_events = [
            e for e in events if "patient" in e.get("event_type", "").lower()
        ]

        metrics = {
            "total_patient_events": len(patient_events),
            "new_patients": len(
                [e for e in patient_events if "created" in e.get("event_type", "")],
            ),
            "patient_appointments": len(
                [e for e in patient_events if "appointment" in e.get("event_type", "")],
            ),
            "patient_assessments": len(
                [e for e in patient_events if "assessment" in e.get("event_type", "")],
            ),
        }

        # Calculate patient engagement
        user_activity = {}
        for event in patient_events:
            user_id = event.get("user_id")
            if user_id:
                user_activity[user_id] = user_activity.get(user_id, 0) + 1

        if user_activity:
            avg_engagement = sum(user_activity.values()) / len(user_activity)
            metrics["avg_patient_engagement"] = round(avg_engagement, 2)
            metrics["highly_engaged_patients"] = len(
                [u for u in user_activity.values() if u > avg_engagement * 1.5],
            )

        return metrics

    def _analyze_therapy_metrics(self, events):
        """Analyze therapy-related metrics."""
        therapy_events = [
            e for e in events if "therapy" in e.get("event_type", "").lower()
        ]

        session_events = [
            e for e in therapy_events if "session" in e.get("event_type", "")
        ]
        completed_sessions = [
            e for e in session_events if "completed" in e.get("event_type", "")
        ]

        return {
            "total_therapy_events": len(therapy_events),
            "therapy_sessions_started": len(
                [e for e in session_events if "started" in e.get("event_type", "")],
            ),
            "therapy_sessions_completed": len(completed_sessions),
            "session_completion_rate": round(
                len(completed_sessions) / max(1, len(session_events)) * 100,
                2,
            ),
            "avg_sessions_per_day": round(
                len(session_events)
                / max(1, (self.config.period_end - self.config.period_start).days),
                2,
            ),
        }

    def _analyze_communication_metrics(self, events):
        """Analyze communication metrics."""
        comm_events = [
            e
            for e in events
            if any(
                term in e.get("event_type", "").lower()
                for term in ["message", "call", "notification"]
            )
        ]

        return {
            "total_communications": len(comm_events),
            "messages_sent": len(
                [e for e in comm_events if "message.sent" in e.get("event_type", "")],
            ),
            "video_calls": len(
                [e for e in comm_events if "video_call" in e.get("event_type", "")],
            ),
            "notifications_sent": len(
                [e for e in comm_events if "notification" in e.get("event_type", "")],
            ),
            "avg_communications_per_day": round(
                len(comm_events)
                / max(1, (self.config.period_end - self.config.period_start).days),
                2,
            ),
        }

    def _analyze_operational_metrics(self, events):
        """Analyze operational efficiency metrics."""
        error_events = [
            e
            for e in events
            if "error" in e.get("event_type", "").lower()
            or "failed" in e.get("event_type", "").lower()
        ]
        auth_events = [e for e in events if "auth" in e.get("event_type", "").lower()]

        return {
            "system_errors": len(error_events),
            "error_rate": round(len(error_events) / max(1, len(events)) * 100, 4),
            "authentication_events": len(auth_events),
            "failed_logins": len(
                [e for e in auth_events if "failed" in e.get("event_type", "")],
            ),
            "system_uptime_score": max(
                0,
                100 - (len(error_events) / max(1, len(events)) * 1000),
            ),
        }

    def _analyze_revenue_impact(self, events):
        """Analyze events that impact revenue."""
        # This would be customized based on your business model
        appointment_events = [
            e for e in events if "appointment" in e.get("event_type", "").lower()
        ]
        completed_appointments = [
            e for e in appointment_events if "completed" in e.get("event_type", "")
        ]

        # Assuming each completed appointment generates revenue
        estimated_revenue_per_appointment = 150  # Configurable

        return {
            "billable_appointments": len(completed_appointments),
            "estimated_revenue": len(completed_appointments)
            * estimated_revenue_per_appointment,
            "appointment_show_rate": round(
                len(completed_appointments) / max(1, len(appointment_events)) * 100,
                2,
            ),
            "revenue_per_day": round(
                (len(completed_appointments) * estimated_revenue_per_appointment)
                / max(1, (self.config.period_end - self.config.period_start).days),
                2,
            ),
        }

    def _generate_recommendations(self, events):
        """Generate business recommendations based on data analysis."""
        recommendations = []

        # Analyze patterns and generate insights
        patient_events = [
            e for e in events if "patient" in e.get("event_type", "").lower()
        ]
        error_events = [e for e in events if "error" in e.get("event_type", "").lower()]

        # Error rate recommendation
        error_rate = len(error_events) / max(1, len(events)) * 100
        if error_rate > 1.0:
            recommendations.append(
                {
                    "type": "system_improvement",
                    "priority": "high",
                    "title": "Reduce System Error Rate",
                    "description": f"Current error rate of {error_rate:.2f}% is above optimal threshold. Consider system optimization.",
                    "impact": "Improved user experience and reduced support burden",
                },
            )

        # Patient engagement recommendation
        user_activity = {}
        for event in patient_events:
            user_id = event.get("user_id")
            if user_id:
                user_activity[user_id] = user_activity.get(user_id, 0) + 1

        if user_activity:
            low_engagement_users = len([u for u in user_activity.values() if u < 3])
            if low_engagement_users > len(user_activity) * 0.3:
                recommendations.append(
                    {
                        "type": "patient_engagement",
                        "priority": "medium",
                        "title": "Improve Patient Engagement",
                        "description": f"{low_engagement_users} patients show low engagement. Consider outreach programs.",
                        "impact": "Better patient outcomes and retention",
                    },
                )

        # More recommendations can be added based on specific business rules

        return recommendations


class ScheduledReportService:
    """Service for managing scheduled report generation."""

    @staticmethod
    def generate_daily_summary():
        """Generate daily summary report."""
        config = ReportConfig(
            title="Daily Analytics Summary",
            description="Comprehensive daily analytics report",
            period_start=timezone.now().replace(
                hour=0,
                minute=0,
                second=0,
                microsecond=0,
            )
            - timedelta(days=1),
            period_end=timezone.now().replace(
                hour=0,
                minute=0,
                second=0,
                microsecond=0,
            ),
            format="pdf",
        )

        generator = AnalyticsSummaryReportGenerator(config)
        return generator.generate()

    @staticmethod
    def generate_weekly_bi_report():
        """Generate weekly business intelligence report."""
        end_date = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        start_date = end_date - timedelta(days=7)

        config = ReportConfig(
            title="Weekly Business Intelligence Report",
            description="Healthcare analytics and business insights",
            period_start=start_date,
            period_end=end_date,
            format="pdf",
        )

        generator = BusinessIntelligenceReportGenerator(config)
        return generator.generate()

    @staticmethod
    def generate_monthly_executive_summary():
        """Generate monthly executive summary."""
        now = timezone.now()
        # First day of current month
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # First day of previous month
        start_date = (start_date - timedelta(days=1)).replace(day=1)
        end_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        config = ReportConfig(
            title="Monthly Executive Summary",
            description="High-level analytics and KPIs for executive review",
            period_start=start_date,
            period_end=end_date,
            format="pdf",
            include_charts=True,
        )

        generator = BusinessIntelligenceReportGenerator(config)
        return generator.generate()
